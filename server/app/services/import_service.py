import io
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.catalog import Product, ProductVariant
from app.models.tenant import Tenant
from app.schemas.catalog_schemas import ProductCreateRequest, ProductVariantSchema, normalize_asset_url
from app.services.catalog_service import _replace_product_images, create_product_service

# One row = one product with a single default variant. Multi-variant-per-row
# import isn't supported in v1 -- add extra variants via the product edit
# screen after import, or a repeated-SKU-with-different-attributes scheme
# later if this becomes a real need.
REQUIRED_COLUMNS = ["name_en", "slug", "base_price", "sku", "stock_quantity"]
TEMPLATE_COLUMNS = [
    "name_en", "name_he", "slug", "description_en", "base_price",
    "sku", "stock_quantity", "category_id", "image_url",
]


def build_import_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_COLUMNS)
    ws.append([
        "Classic T-Shirt", "חולצה קלאסית", "classic-tshirt", "100% cotton",
        25.00, "TSHIRT-RED-M", 50, "", "",
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_products_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read file -- expected a valid .xlsx spreadsheet",
        )

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Spreadsheet is empty")

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column(s): {', '.join(missing)}",
        )

    # Last column wins on a duplicate header rather than erroring -- keeps
    # the parser simple, and a sensibly-built sheet won't hit this anyway.
    col_index = {name: idx for idx, name in enumerate(header) if name}

    parsed_rows: List[Dict[str, Any]] = []
    for row_number, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(v is None for v in raw_row):
            continue
        data = {name: (raw_row[idx] if idx < len(raw_row) else None) for name, idx in col_index.items()}
        parsed_rows.append({"row_number": row_number, "data": data, "errors": _validate_row(data)})

    return {
        "rows": parsed_rows,
        "valid_count": sum(1 for r in parsed_rows if not r["errors"]),
        "total_count": len(parsed_rows),
    }


def _validate_row(data: Dict[str, Any]) -> List[str]:
    # name_en/slug are only required for a genuinely new SKU -- that can't be
    # decided here (no DB access in this pure parsing/validation step), so
    # it's enforced in commit_products_import's create branch instead, once
    # we actually know the SKU doesn't exist yet. A row that's only updating
    # an existing SKU's stock/price is valid without them.
    errors: List[str] = []
    if not data.get("sku"):
        errors.append("sku is required")

    try:
        if Decimal(str(data.get("base_price"))) <= 0:
            errors.append("base_price must be greater than 0")
    except (InvalidOperation, TypeError):
        errors.append("base_price must be a number")

    try:
        if int(data.get("stock_quantity")) < 0:
            errors.append("stock_quantity cannot be negative")
    except (TypeError, ValueError):
        errors.append("stock_quantity must be a whole number")

    return errors


async def commit_products_import(tenant_slug: str, rows: List[Dict[str, Any]], db: AsyncSession) -> Dict[str, Any]:
    tenant_result = await db.execute(select(Tenant.id).where(Tenant.slug == tenant_slug))
    tenant_id = tenant_result.scalar_one_or_none()
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tenant not found")

    created: List[Dict[str, Any]] = []
    updated: List[Dict[str, Any]] = []
    failed: List[Dict[str, Any]] = []

    for row in rows:
        data = row.get("data", row)
        row_number = row.get("row_number")
        sku = str(data.get("sku") or "").strip()
        errors = row.get("errors") or _validate_row(data)
        if errors:
            failed.append({"row_number": row_number, "sku": sku or None, "error": "; ".join(errors)})
            continue

        try:
            # Upsert by SKU (scoped to this tenant): a spreadsheet re-uploaded
            # with updated stock numbers updates existing products instead of
            # erroring on a duplicate slug -- this is what makes one import
            # path cover both "add new products" and "update inventory".
            variant_result = await db.execute(
                select(ProductVariant).where(ProductVariant.tenant_id == tenant_id, ProductVariant.sku == sku)
            )
            existing_variant = variant_result.scalar_one_or_none()

            if existing_variant:
                existing_variant.stock_quantity = int(data["stock_quantity"])
                existing_variant.price_override = Decimal(str(data["base_price"]))
                image_url_raw = data.get("image_url")
                image_url = str(image_url_raw).strip() if image_url_raw else ""
                if image_url:
                    image_url = normalize_asset_url(image_url, field_name="image_url")

                name_en = str(data["name_en"]).strip() if data.get("name_en") else ""
                name_he = str(data["name_he"]).strip() if data.get("name_he") else ""
                description_en = str(data["description_en"]).strip() if data.get("description_en") else ""
                category_raw = data.get("category_id")
                has_category = category_raw is not None and str(category_raw).strip() != ""

                needs_product = bool(image_url or name_en or name_he or description_en or has_category)
                product = None
                if needs_product:
                    product_result = await db.execute(
                        select(Product)
                        .where(Product.id == existing_variant.product_id, Product.tenant_id == tenant_id)
                        .options(selectinload(Product.images))
                    )
                    product = product_result.scalar_one()
                    if name_en or name_he:
                        current_name = dict(product.name or {})
                        if name_en:
                            current_name["en"] = name_en
                        if name_he:
                            current_name["he"] = name_he
                        product.name = current_name
                    if description_en:
                        current_desc = dict(product.description or {})
                        current_desc["en"] = description_en
                        product.description = current_desc
                    if has_category:
                        product.category_id = int(category_raw)
                    if image_url:
                        await _replace_product_images(db, tenant_id, product, [image_url])
                await db.commit()
                updated_row: Dict[str, Any] = {
                    "row_number": row_number,
                    "sku": sku,
                    "variant_id": existing_variant.id,
                    "product_id": existing_variant.product_id,
                    "stock_quantity": int(data["stock_quantity"]),
                    "base_price": str(Decimal(str(data["base_price"]))),
                }
                if product is not None:
                    updated_row["name"] = product.name
                    updated_row["description"] = product.description
                    updated_row["category_id"] = product.category_id
                if image_url:
                    updated_row["primary_image_url"] = image_url
                    updated_row["images"] = [image_url]
                updated.append(updated_row)
            else:
                if not data.get("name_en") or not data.get("slug"):
                    raise ValueError("name_en and slug are required to create a new product (sku not found)")
                # Matches the manual "New Product" form's fallback (products/new/page.tsx):
                # a store can require translations for languages beyond English, and a
                # spreadsheet with a single English column shouldn't fail every row over
                # missing Hebrew text -- fall back to the English value like the form does.
                name_en = str(data["name_en"])
                description_en = str(data["description_en"]) if data.get("description_en") else None
                create_image_url = str(data["image_url"]).strip() if data.get("image_url") else ""
                if create_image_url:
                    create_image_url = normalize_asset_url(create_image_url, field_name="image_url")
                req = ProductCreateRequest(
                    category_id=int(data["category_id"]) if data.get("category_id") else None,
                    name={"en": name_en, "he": str(data["name_he"]) if data.get("name_he") else name_en},
                    slug=str(data["slug"]),
                    description={"en": description_en, "he": description_en} if description_en else None,
                    base_price=Decimal(str(data["base_price"])),
                    variants=[ProductVariantSchema(sku=sku, stock_quantity=int(data["stock_quantity"]))],
                    images=[create_image_url] if create_image_url else [],
                )
                product = await create_product_service(tenant_slug, req, db)
                created.append({"row_number": row_number, "sku": sku, "product_id": product.id})
        except HTTPException as e:
            await db.rollback()
            failed.append({"row_number": row_number, "sku": sku or None, "error": str(e.detail)})
        except Exception as e:
            await db.rollback()
            failed.append({"row_number": row_number, "sku": sku or None, "error": str(e)})

    return {
        "created_count": len(created), "updated_count": len(updated), "failed_count": len(failed),
        "created": created, "updated": updated, "failed": failed,
    }
